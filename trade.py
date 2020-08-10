# noinspection PyUnresolvedReferences
import pandas as pd
import requests
from openpyxl.utils import get_column_letter
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time
import dateutil
import json

pd.set_option('display.width', 1500)
pd.set_option('display.max_rows', 150)
pd.set_option('display.max_columns', 50)
excel_out_file = r'C:\Users\raj\Desktop\stock.xlsx'
excel_read_file = r'C:\Users\raj\Desktop\my trade.xls'
sheet1 = 'Portfolio'
sheet2 = 'sold stock'
sheet3 = 'NSEDATA'
sheet4 = 'UPSTOX_DATA'


def read_excel_file_calc():
    df = pd.read_excel(excel_read_file, sheet_name='Output')
    df = df[['Company', 'Side', 'Qty', 'Price']]
    df['Total'] = df['Qty'] * df['Price']
    # df = df.set_index(['Company', 'Side'])
    df = df.groupby(['Company', 'Side'])[['Qty', 'Total']].apply(sum).unstack().reset_index()

    df = df.to_numpy()
    df = pd.DataFrame(df)
    df = df.fillna(0)

    df = df.rename(columns={0: 'Company Name', 1: 'Qty_Buy', 2: 'Qty_Sell', 3: 'Total_Buy_Amt', 4: 'Total_Sell_Amt'})
    df = df.sort_values(by=['Total_Sell_Amt'], ascending=True)
    df['Net_Quentity'] = df['Qty_Buy'] - df['Qty_Sell']
    df['Net_Quentity'].astype(int)
    df = df.sort_values(by=['Net_Quentity'], ascending=False)
    df['Net_Amount'] = df['Total_Buy_Amt'] - df['Total_Sell_Amt']
    df['Buy_Average'] = df['Net_Amount'].div(df['Net_Quentity'].where(df['Net_Quentity'] != 0, np.nan))
    df = df.fillna(0)
    df['Per'] = df['Net_Quentity'] * df['Buy_Average']
    df['Percent'] = np.where(df['Net_Quentity'] != 0, df['Net_Amount'] * 100 / df['Per'].sum(), 0)
    df = df.drop(columns=['Per'])
    df = df.sort_values(by=['Percent'], ascending=False)
    df['Percent'].astype(int)

    df1 = df[df['Net_Quentity'] != 0]
    df2 = df[df['Net_Quentity'] <= 0]
    df2 = df2.drop(columns=['Net_Quentity', 'Buy_Average', 'Percent'])
    return df1, df2


def data_get(url, headers, cookies):
    session = requests.session()
    # for cookie in cookie_dict:
    #     session.cookies.set(cookie, cookie_dict[cookie])
    dict = session.get(url, headers=headers).json()
    # r = requests.get(url=url)
    return dict


def convert_rupee(val):
    new_val = val.replace(",", "")
    return float(new_val)


def main(dict):
    df = pd.DataFrame(dict['data'])
    df = df[['symbol', 'open', 'high', 'low', 'ltP', 'ptsC', 'per', 'trdVol', 'wkhi', 'wklo']]
    # df.set_index('symbol', inplace = True)
    df['open'] = df['open'].apply(convert_rupee)
    df['high'] = df['high'].apply(convert_rupee)
    df['low'] = df['low'].apply(convert_rupee)
    df['ltP'] = df['ltP'].apply(convert_rupee)
    df['ptsC'] = df['ptsC'].apply(convert_rupee)
    df['per'] = df['per'].apply(convert_rupee)
    df['trdVol'] = df['trdVol'].apply(convert_rupee)
    df['wkhi'] = df['wkhi'].apply(convert_rupee)
    df['wklo'] = df['wklo'].apply(convert_rupee)
    df = df.rename(
        columns={'open': 'Open', 'high': 'High', 'low': 'Low', 'ltp': 'LTP', 'ptsC': 'Change', 'per': '%Change',
                 'trdVol': 'Volume', 'wkhi': '52w H', 'wklo': '52w L'})
    return df


def as_text(value):
    if value is None:
        return ""
    return str(value)


def populate_sheet(df, ws):
    dfrow = 1
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
        dfrow += 1
    title_row = '1'
    value_cells = 'B2:{col}{row}'.format(
        col=get_column_letter(ws.max_column),
        row=ws.max_row)
    index_column = 'A'

    # for general styling, one has to iterate over all cells individually
    for row in ws[value_cells]:
        for cell in row:
            cell.style = 'Normal'
            cell.number_format = '0.00'
    # builtin or named styles can be applied by using the object or their name
    # https://openpyxl.readthedocs.io/en/stable/styles.html#using-builtin-styles
    for cell in ws[index_column]:
        cell.style = 'Normal'

    # style header line last, so that headline style wins in cell A1
    for cell in ws[title_row]:
        # cell.style = 'Headline 2'
        cell.style = 'Accent2'

    return ws, dfrow


def cell_width_alignment(ws):
    for column_cells in ws.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 3


def add_nse_data(port_df, nse_df):
    df_mapping = pd.read_excel(excel_read_file, sheet_name='name_mapping')
    port_df = port_df.merge(df_mapping, left_on='Company Name', right_on='company_name', how='left').drop(
        columns=['company_name'])
    port_df = port_df.merge(nse_df, left_on='nse_name', right_on='symbol', how='left').drop(
        columns=['nse_name', 'symbol'])

    port_df = port_df.fillna(0)
    # sum1 = port_df['Net_Amount'].sum()
    # port_df = port_df.append([{'Net_Quentity': 'Total Invest Rs.', 'Net_Amount': sum1}], ignore_index=True)
    port_df = round(port_df, 2)
    return port_df


if __name__ == "__main__":
    # url_midcap50 = 'https://www1.nseindia.com/live_market/dynaContent/live_watch/stock_watch/niftyMidcap50StockWatch.json'
    # url_juniornifty = 'https://www1.nseindia.com/live_market/dynaContent/live_watch/stock_watch/juniorNiftyStockWatch.json'
    url = 'https://www1.nseindia.com/live_market/dynaContent/live_watch/stock_watch/nifty500StockWatch.json'
    # url_nifty50 = 'https://www1.nseindia.com/live_market/dynaContent/live_watch/stock_watch/niftyStockWatch.json'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.89 Safari/537.36',
        "Accept-Language": 'en-US,en;q=0.9', "Accept-Encoding": 'gzip, deflate'}
    cookie_dict = {
        'bm_sv': 'bm_sv=B683127B319CDEE635D5372C90911E65~9bcie0JgYimO/ip/zZyr7MogxfyXlHq+Tz5Ui7Zhe2uEabg2yRXR4tGEB6fLuPo5NOfwvNh+fLoL+24U2NS/6RnomTLCaKkqrvMGVymRDeXQvV0BPqISClZsOss1CDEbSSLSCr0PBEZlCovszNGVtGObdpFqxB7xlKx'}

    # df_nifty50 = main(data_get(url_nifty50, headers, cookie_dict))
    # df_midcap50 = main(data_get(url_midcap50, headers, cookie_dict))
    # df_juniornifty = main(data_get(url_juniornifty, headers, cookie_dict))
    df_row = main(data_get(url, headers, cookie_dict))
    # df_nifty50['M_Cap'] = 'Nifty 50'
    # df_midcap50['M_cap'] = 'Midcap50'
    # df_juniornifty['M_Cap'] = 'Nifty Ju'
    # df1 = df[df['per'] < -1]

    # df_row = pd.concat([df_nifty50, df_juniornifty, df_midcap50], ignore_index=True).reset_index()
    # df_row = pd.merge(df_row, df_stock_watch, on='symbol', how='outer')
    # # df_row = df_row.merge(df_stock_watch, left_on='symbol', right_on='symbol', how='left')
    # df_row = df_row.drop(columns=['index'])
    df_row = df_row.sort_values(by=['%Change'], ascending=False)
    df1, df2 = read_excel_file_calc()
    df1 = add_nse_data(df1, df_row)
    df1 = round(df1, 2)
    df2 = round(df2, 2)
    df_row = round(df_row, 2)
    wb = Workbook()
    # When you make a new workbook you get a new blank active sheet
    # We need to delete it since we do not want it
    wb.remove(wb.active)

    sheet1 = wb.create_sheet(title=sheet1)
    sheet2 = wb.create_sheet(title=sheet2)
    sheet3 = wb.create_sheet(title=sheet3)
    # sheet4 = wb.create_sheet(title=sheet4)
    sheet1.sheet_properties.tabColor = "1072BA"
    sheet2.sheet_properties.tabColor = "1072BA"
    sheet3.sheet_properties.tabColor = "1072BA"
    # sheet4.sheet_properties.tabColor = "1072BA"

    ws1, dfrow1 = populate_sheet(df1, sheet1)
    sum1 = df1['Net_Amount'].sum()
    sum1 = round(sum1, 2)
    ws1.cell(row=dfrow1 + 1, column=1, value='Total Investment Rs.').style = 'Total'
    ws1.cell(row=dfrow1 + 1, column=2, value=sum1).style = 'Total'
    cell_width_alignment(ws1)

    ws2, dfrow2 = populate_sheet(df2, sheet2)
    sum2 = df2['Net_Amount'].sum()
    sum2 = round(sum2, 2)
    if sum2 > 0:
        ws2.cell(row=dfrow2 + 1, column=1, value='Total Loss Rs.').style = 'Total'
        ws2.cell(row=dfrow2 + 1, column=2, value=sum2).style = 'Total'
    else:
        ws2.cell(row=dfrow2 + 1, column=1, value='Total Profit Rs.').style = 'Total'
        ws2.cell(row=dfrow2 + 1, column=2, value=sum2).style = 'Total'
    # df2 = df2.append([{'Net_Quentity': 'Total Loss/Profit Rs.', 'Net_Amount': sum2}], ignore_index=True)

    cell_width_alignment(ws2)
    ws3, dfrow3 = populate_sheet(df_row, sheet3)
    cell_width_alignment(ws3)
    # ws4, dfrow4 = populate_sheet(df3, sheet4)
    # cell_width_alignment(ws4)

    wb.save(excel_out_file)
