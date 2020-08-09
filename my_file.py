'''
Video Tutorial: https://youtu.be/1Kcco6koC34
Description: Better Techniques to automate excel reporting using pandas and python.
'''
import pandas as pd
import numpy as np
from io import StringIO
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Border, Side, PatternFill, Font, Color, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
# output_filename = __file__.replace(".py", ".xlsx")
import matplotlib.pyplot as plt
pd.set_option('display.width', 1500)
pd.set_option('display.max_rows', 1500)
pd.set_option('display.max_columns', 50)

excel_file = 'stock.xlsx'
def read_excel_file_calc():
    df = pd.read_excel(excel_file, sheet_name='Sheet1')
    df = df[['Company', 'Side', 'Qty', 'Price']]
    df['Total'] = df['Qty'] * df['Price']
    # df = df.set_index(['Company', 'Side'])
    df = df.groupby(['Company', 'Side'])[['Qty','Total']].apply(sum).unstack().reset_index()

    df = df.to_numpy()
    df = pd.DataFrame(df)
    df = df.fillna(0)

    df = df.rename(columns={0:'Company Name', 1:'Qty_Buy', 2:'Qty_Sell', 3:'Total_Buy_Amt', 4:'Total_Sell_Amt'})
    df = df.sort_values(by=['Total_Sell_Amt'], ascending=True)
    df['Net_Quentity'] = df['Qty_Buy'] - df['Qty_Sell']
    df['Net_Quentity'].astype(int)
    df = df.sort_values(by=['Net_Quentity'], ascending=False)
    df['Net_Amount'] = df['Total_Buy_Amt'] - df['Total_Sell_Amt']
    df['Buy_Average'] = df['Net_Amount'].div(df['Net_Quentity'].where(df['Net_Quentity'] != 0, np.nan))
    df = df.fillna(0)
    df['Per'] = df['Net_Quentity']*df['Buy_Average']
    df['Percent'] = np.where(df['Net_Quentity']!= 0, df['Net_Amount']*100 / df['Per'].sum(), 0)
    df = df.drop(columns=['Per'])
    df = df.sort_values(by=['Percent'], ascending=False)
    df['Percent'].astype(int)

    df1 = df[df['Net_Quentity']!= 0 ]
    sum1 = df1['Net_Amount'].sum()
    df1 = round(df1, 2)
    df2 = df[df['Net_Quentity'] <= 0 ]
    sum2 = df2['Net_Amount'].sum()
    df2 = round(df2, 2)
    df1 = df1.append([{'Net_Quentity':'Total Invest Rs.','Net_Amount':sum1}], ignore_index=True)
    df2 = df2.append([{'Net_Quentity':'Total Loss/Profit Rs.','Net_Amount':sum2}], ignore_index=True)
    return df1, df2

def as_text(value):
    if value is None:
        return ""
    return str(value)

def portfolio_stock(sheet_name, df):
    # Unpacking arguments using '**' for dictionary of keyword arguments
    # see: https://docs.python.org/3/tutorial/controlflow.html#more-on-defining-functions
    writer_args = {
        'path': excel_file,
        'mode': 'a',
        'engine': 'openpyxl'}


    with pd.ExcelWriter(**writer_args) as xlsx:
        df.to_excel(xlsx, sheet_name, index=False)

        # worksheets that have been created with this ExcelWriter can be accessed
        # by openpyxl using its API. `ws` is now a openpyxl Worksheet object
        wb = xlsx.book
        ws = xlsx.sheets[sheet_name]

        # cell ranges
        title_row = '1'
        value_cells = 'B2:{col}{row}'.format(
            col=get_column_letter(ws.max_column),
            row=ws.max_row)
        index_column = 'A'


        # for general styling, one has to iterate over all cells individually
        for row in ws[value_cells]:
            for cell in row:
                cell.style = 'Value Cell'
                cell.number_format = '0.00'
        # builtin or named styles can be applied by using the object or their name
        # https://openpyxl.readthedocs.io/en/stable/styles.html#using-builtin-styles
        for cell in ws[index_column]:
            cell.style = 'Index Style'

        # style header line last, so that headline style wins in cell A1
        for cell in ws[title_row]:
            # cell.style = 'Headline 2'
            cell.style = 'Accent2'
        for column_cells in ws.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length*1.2



if __name__ == "__main__":

    sheet1 = 'Portfolio'
    sheet2 = 'sold stock'
    df1,df2 = read_excel_file_calc()
    portfolio_stock(sheet1, df1)
    portfolio_stock(sheet2, df2)

