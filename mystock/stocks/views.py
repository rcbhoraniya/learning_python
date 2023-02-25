import csv
import io
import pandas as pd
import numpy as np
import requests as r
import openpyxl

from django.db.models.functions import Cast
# from django.http import HttpResponseRedirect
# from django.urls import reverse_lazy
from django.http import HttpResponse, FileResponse
from django.views.generic import *
# from .models import *
from django.db.models import *
from datetime import *
from django.db.models import Subquery
from .forms import *
from django.contrib.auth.mixins import PermissionRequiredMixin, LoginRequiredMixin
from django.contrib.auth.views import *
from django.contrib.auth.forms import *
from django.shortcuts import redirect, get_object_or_404
import yfinance as yf
from datetime import datetime, timedelta, date, time
import xlsxwriter
from django.utils import timezone

TODAY_DATETIME = datetime.now()
TODAY_DATE = datetime.now().date()
TODAY_S = TODAY_DATE.strftime('%Y-%m-%d') + ' 15:31:00'
MARKET_CLOSE_DATETIME = pd.to_datetime(TODAY_S)

pd.set_option('display.width', 1500)
pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 50)

proxy_host = "10.72.44.45"
proxy_port = "808"
proxy_auth = "wtrmrb:wtrmrb"
PROXIES = {
    "https": "https://{}@{}:{}/".format(proxy_auth, proxy_host, proxy_port),
    "http": "http://{}@{}:{}/".format(proxy_auth, proxy_host, proxy_port)
}


class PermissionDeniedView(TemplateView):
    template_name = 'stocks/permission_denied.html'


class UserAccessMixin(PermissionRequiredMixin, LoginRequiredMixin):

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path(), self.get_login_url(), self.get_redirect_field_name())
        if not self.has_permission():
            return redirect('stocks:permissiondenied')
        return super(UserAccessMixin, self).dispatch(request, *args, **kwargs)


class UserRegistrationView(CreateView):
    model = User
    form_class = UserRegistrationForm
    template_name = 'stocks/registration.html'
    success_url = reverse_lazy('stocks:login')


class UserLoginView(LoginView):
    form_class = LoginForm
    template_name = 'stocks/login.html'
    next_page = 'stocks:index'
    # redirect_field_name = 'stocks:index'


class UserLogoutView(LogoutView):
    # template_name = 'stocks/logout.html'
    next_page = 'stocks:login'


class IndexView(UserAccessMixin, ListView):
    permission_required = ('stocks.view_stockmap',)

    template_name = 'portfolio/index.html'
    context_object_name = 'stock_list'

    # paginate_by = 20

    def get_queryset(self):
        stockdata_subqs = StockData.objects.filter(company=OuterRef("pk")).filter(
            company__is_portfolio_stock=True).values(
            'company').annotate(qty_sum=Sum(F('quantity') * F('side')),
                                               price_sum=Sum(F('price') * F('side') * F('quantity')),
                                               average_price=Sum(F('price') * F('side') * F('quantity')) / Sum(
                                                   F('quantity') * F('side')))
        nsedata_subqs = NSEBhavcopy.objects.filter(symbol=OuterRef("pk"))

        data = StockMap.objects.filter(is_portfolio_stock=True).annotate(
            qty_sum=Subquery(stockdata_subqs.values('qty_sum')),
            price_sum=Subquery(stockdata_subqs.values('price_sum')[:1]),
            average_price=Subquery(stockdata_subqs.values('average_price')[:1]),
            close=Subquery(nsedata_subqs.values('close')[:1]),
            date=Subquery(nsedata_subqs.values('date')[:1]),
            w52_high=Subquery(nsedata_subqs.values('w52_high')[:1]),
            w52_low=Subquery(nsedata_subqs.values('w52_low')[:1]),
            change=Subquery(nsedata_subqs.values('change')[:1]),
            pchange=Subquery(nsedata_subqs.values('perchange')[:1]),
            per_down=F('close') * 100 / F('w52_high') - 100,
            today_value=F('close') * F('qty_sum'),
            profit=F('today_value') - F('price_sum')
        ).order_by('name')
        # for item in data:
        #     print(item.name,item.profit)
        # print(data)
        total_investment = 0
        profit_sum = 0

        for item in data:
            total_investment += item.price_sum
            profit_sum += item.profit

        data = data.annotate(
            total_investment=Cast(total_investment, output_field=DecimalField(max_digits=12, decimal_places=2)),
            profit_sum=Cast(profit_sum, output_field=DecimalField(max_digits=12, decimal_places=2)))
        data = data.annotate(portweight=F('price_sum') * 100 / F('total_investment'),
                             profit_percentage=F('today_value') * 100 / F('price_sum') - 100,
                             port_per_up=F('profit_sum') * 100 / F('total_investment'))
        # print(data[0].id,data[0].name)
        return data


class SectorStockListView(UserAccessMixin, ListView):
    permission_required = ('stocks.view_stockmap',)

    template_name = 'portfolio/index.html'
    context_object_name = 'stock_list'

    def get_queryset(self):
        print(self.kwargs['sector'])
        stockdata_subqs = StockData.objects.filter(company=OuterRef("pk")).filter(
            company__sector__sector=self.kwargs['sector']).values(
            'company').annotate(qty_sum=Sum(F('quantity') * F('side')),
                                price_sum=Sum(F('price') * F('side') * F('quantity')),
                                average_price=Sum(F('price') * F('side') * F('quantity')) / Sum(
                                    F('quantity') * F('side')))
        nsedata_subqs = NSEBhavcopy.objects.filter(symbol=OuterRef("pk"))

        data = StockMap.objects.filter(is_portfolio_stock=True).filter(sector__sector=self.kwargs['sector']).annotate(
            qty_sum=Subquery(stockdata_subqs.values('qty_sum')),
            price_sum=Subquery(stockdata_subqs.values('price_sum')[:1]),
            average_price=Subquery(stockdata_subqs.values('average_price')[:1]),
            close=Subquery(nsedata_subqs.values('close')[:1]),
            date=Subquery(nsedata_subqs.values('date')[:1]),
            w52_high=Subquery(nsedata_subqs.values('w52_high')[:1]),
            w52_low=Subquery(nsedata_subqs.values('w52_low')[:1]),
            change=Subquery(nsedata_subqs.values('change')[:1]),
            pchange=Subquery(nsedata_subqs.values('perchange')[:1]),
            per_down=F('close') * 100 / F('w52_high') - 100,
            today_value=F('close') * F('qty_sum'),
            profit=F('today_value') - F('price_sum')
        ).order_by('name')
        data = data.filter(qty_sum__gt=0)
        print(data)

        for item in data:
            print(item.name, item.profit)

        total_investment = 0
        profit_sum = 0

        for item in data:
            total_investment += item.price_sum
            profit_sum += item.profit

        data = data.annotate(
            total_investment=Cast(total_investment, output_field=DecimalField(max_digits=12, decimal_places=2)),
            profit_sum=Cast(profit_sum, output_field=DecimalField(max_digits=12, decimal_places=2)))
        data = data.annotate(portweight=F('price_sum') * 100 / F('total_investment'),
                             profit_percentage=F('today_value') * 100 / F('price_sum') - 100,
                             port_per_up=F('profit_sum') * 100 / F('total_investment'))
        # print(data)
        return data


class StockDetailView(UserAccessMixin, ListView):
    permission_required = ('stocks.view_stockmap',)

    model = StockMap
    template_name = 'stocks/portfolio_stock_detail.html'
    context_object_name = 'stock_list'
    year = 1

    def get_queryset(self, ):
        search_text = self.request.GET.get('search')
        if search_text == '1_year_chart':
            self.year = 1
        elif search_text == '2_year_chart':
            self.year = 2
        elif search_text == '3_year_chart':
            self.year = 3
        elif search_text == '4_year_chart':
            self.year = 4
        elif search_text == '5_year_chart':
            self.year = 5
        elif search_text == '10_year_chart':
            self.year = 10

        start_date = TODAY_DATE - timedelta(days=365 * self.year)

        self.stockmap = get_object_or_404(StockMap, id=self.kwargs['pk'])
        nse_data = NSEBhavcopy.objects.filter(symbol=self.stockmap).first()
        # print(nse_data.date)
        hist_data = HistoricalData.objects.filter(Q(company=self.stockmap) & Q(date__gte=start_date)).order_by('date')
        year_avg = hist_data.aggregate(Avg('close'))
        # print(year_avg['close__avg'])
        historical_data = HistoricalData.objects.filter(Q(company=self.stockmap) & Q(date__gte=start_date)).order_by(
            'date').annotate(
            avg_200=Window(expression=Avg('close'), order_by=F('date').asc(), frame=RowRange(start=-199, end=0)),
            avg_100=Window(expression=Avg('close'), order_by=F('date').asc(), frame=RowRange(start=-99, end=0)),
            avg_50=Window(expression=Avg('close'), order_by=F('date').asc(), frame=RowRange(start=-49, end=0)))
        # print(historical_data)
        historical_data_last_obj = historical_data.order_by('date').last()
        # Fibonacci Pivot Points
        # Pivot Point(P) = (High + Low + Close) / 3
        # Support 1(S1) = P - {.382 * (High - Low)}
        # Support 2(S2) = P - {.618 * (High - Low)}
        # Support 3(S3) = P - {1 * (High - Low)}
        # Resistance 1(R1) = P + {.382 * (High - Low)}
        # Resistance 2(R2) = P + {.618 * (High - Low)}
        # Resistance 3(R3) = P + {1 * (High - Low)}
        pivot = (historical_data_last_obj.close + historical_data_last_obj.low + historical_data_last_obj.high) / 3
        h_l = historical_data_last_obj.high - historical_data_last_obj.low
        s1 = pivot - h_l * 382 / 1000
        s2 = pivot - h_l * 618 / 1000
        s3 = pivot - h_l * 1
        r1 = pivot + h_l * 382 / 1000
        r2 = pivot + h_l * 618 / 1000
        r3 = pivot + h_l * 1

        maximum_price = historical_data.order_by('-high').first()
        minimum_price = historical_data.order_by('low').first()
        difference = maximum_price.high - minimum_price.low
        first_level = maximum_price.high - difference * 236 / 1000
        second_level = maximum_price.high - difference * 382 / 1000
        third_level = maximum_price.high - difference * 5 / 10
        fourth_level = maximum_price.high - difference * 618 / 1000

        queryset = StockData.objects.filter(company=self.stockmap).values(
            'company__name', 'company').annotate(
            qty_sum=Sum(F('quantity') * F('side')),
            price_sum=Sum(F('price') * F('side') * F('quantity')),
            average_price=Sum(F('price') * F('side') * F('quantity')) / Sum(F('quantity') * F('side')),
            today_value=Sum(F('quantity') * F('side') * nse_data.close),
            profit=Sum(F('quantity') * F('side')) * nse_data.close - Sum(F('price') * F('side') * F('quantity')),
            close=Value(nse_data.close, output_field=DecimalField(max_digits=12, decimal_places=2)),
            date=Value(nse_data.date, output_field=DateTimeField()),
            w52_high=Value(nse_data.w52_high, output_field=DecimalField(max_digits=12, decimal_places=2)),
            w52_low=Value(nse_data.w52_low, output_field=DecimalField(max_digits=12, decimal_places=2)),
            per_down=F('close') * 100 / F('w52_high') - 100,
            w52_avg=Value(year_avg['close__avg'], output_field=DecimalField(max_digits=12, decimal_places=2)),
            profit_percentage=F('today_value') * 100 / F('price_sum') - 100,
            # corporate_action=Value(nse_data.corporate_action, output_field=CharField(max_length=100)),
            # x_date=Value(nse_data.x_date, output_field=DateTimeField()),
            year_high=Value(maximum_price.high, output_field=DecimalField(max_digits=12, decimal_places=2)),
            year_high_date=Value(maximum_price.date, output_field=DateTimeField()),
            year_low=Value(minimum_price.low, output_field=DecimalField(max_digits=12, decimal_places=2)),
            year_low_date=Value(minimum_price.date, output_field=DateTimeField()),
            first_level=Value(first_level, output_field=DecimalField(max_digits=12, decimal_places=2)),
            second_level=Value(second_level, output_field=DecimalField(max_digits=12, decimal_places=2)),
            third_level=Value(third_level, output_field=DecimalField(max_digits=12, decimal_places=2)),
            fourth_level=Value(fourth_level, output_field=DecimalField(max_digits=12, decimal_places=2)),
            avg_200=Value(historical_data_last_obj.avg_200, output_field=DecimalField(max_digits=12, decimal_places=2)),
            avg_100=Value(historical_data_last_obj.avg_100, output_field=DecimalField(max_digits=12, decimal_places=2)),
            avg_50=Value(historical_data_last_obj.avg_50, output_field=DecimalField(max_digits=12, decimal_places=2)),
            s1=Value(s1, output_field=DecimalField(max_digits=12, decimal_places=2)),
            s2=Value(s2, output_field=DecimalField(max_digits=12, decimal_places=2)),
            s3=Value(s3, output_field=DecimalField(max_digits=12, decimal_places=2)),
            r1=Value(r1, output_field=DecimalField(max_digits=12, decimal_places=2)),
            r2=Value(r2, output_field=DecimalField(max_digits=12, decimal_places=2)),
            r3=Value(r3, output_field=DecimalField(max_digits=12, decimal_places=2)),
            pivot=Value(pivot, output_field=DecimalField(max_digits=12, decimal_places=2)),

        )
        # print(queryset)
        return queryset

    def get_context_data(self, **kwargs):
        end_date = TODAY_DATE - timedelta(days=365 * self.year)
        # Call the base implementation first to get a context
        context = super().get_context_data(**kwargs)

        queryset = HistoricalData.objects.filter(
            Q(company=self.kwargs['pk']) & Q(date__range=(end_date, TODAY_DATE))).order_by(
            'date')
        maximum_price = queryset.order_by('-high').first()
        minimum_price = queryset.order_by('low').first()
        difference = maximum_price.high - minimum_price.low
        first_level = maximum_price.high - difference * 236 / 1000
        second_level = maximum_price.high - difference * 382 / 1000
        third_level = maximum_price.high - difference * 5 / 10
        fourth_level = maximum_price.high - difference * 618 / 1000
        data = queryset.annotate(
            year_high=Value(maximum_price.high, output_field=DecimalField(max_digits=12, decimal_places=2)),
            year_high_date=Value(maximum_price.date, output_field=DateTimeField()),
            year_low=Value(minimum_price.low, output_field=DecimalField(max_digits=12, decimal_places=2)),
            year_low_date=Value(minimum_price.date, output_field=DateTimeField()),
            first_level=Value(first_level, output_field=DecimalField(max_digits=12, decimal_places=2)),
            second_level=Value(second_level, output_field=DecimalField(max_digits=12, decimal_places=2)),
            third_level=Value(third_level, output_field=DecimalField(max_digits=12, decimal_places=2)),
            fourth_level=Value(fourth_level, output_field=DecimalField(max_digits=12, decimal_places=2)),
            avg_200=Window(expression=Avg('close'), order_by=F('date').asc(), frame=RowRange(start=-199, end=0)),
            avg_100=Window(expression=Avg('close'), order_by=F('date').asc(), frame=RowRange(start=-99, end=0)),
            avg_50=Window(expression=Avg('close'), order_by=F('date').asc(), frame=RowRange(start=-49, end=0)),
        ).order_by('date')

        context['stock_hist'] = data
        context['stockdata_list'] = StockMap.objects.get(id=self.kwargs['pk']).stockdata_set.all().order_by('date')

        return context


class HistoricalDataView(UserAccessMixin, ListView):
    permission_required = ('stocks.view_stockmap',)

    template_name = 'stocks/portfolio_stock_history.html'
    context_object_name = 'stock_hist'
    year = 1

    # paginate_by = 20

    def get_queryset(self):

        search_text = self.request.GET.get('search')
        if search_text == '1_year_chart':
            self.year = 1
        elif search_text == '2_year_chart':
            self.year = 2
        elif search_text == '3_year_chart':
            self.year = 3
        elif search_text == '4_year_chart':
            self.year = 4
        elif search_text == '5_year_chart':
            self.year = 5
        elif search_text == '10_year_chart':
            self.year = 10

        end_date = TODAY_DATE - timedelta(days=365 * self.year)
        queryset = HistoricalData.objects.filter(
            Q(company=self.kwargs['pk']) & Q(date__range=(end_date, TODAY_DATE))).order_by(
            'date')
        maximum_price = queryset.order_by('-high').first()
        minimum_price = queryset.order_by('low').first()
        difference = maximum_price.high - minimum_price.low
        first_level = maximum_price.high - difference * 236 / 1000
        second_level = maximum_price.high - difference * 382 / 1000
        third_level = maximum_price.high - difference * 5 / 10
        fourth_level = maximum_price.high - difference * 618 / 1000
        data = queryset.annotate(
            year_high=Value(maximum_price.high, output_field=DecimalField(max_digits=12, decimal_places=2)),
            year_high_date=Value(maximum_price.date, output_field=DateTimeField()),
            year_low=Value(minimum_price.low, output_field=DecimalField(max_digits=12, decimal_places=2)),
            year_low_date=Value(minimum_price.date, output_field=DateTimeField()),
            first_level=Value(first_level, output_field=DecimalField(max_digits=12, decimal_places=2)),
            second_level=Value(second_level, output_field=DecimalField(max_digits=12, decimal_places=2)),
            third_level=Value(third_level, output_field=DecimalField(max_digits=12, decimal_places=2)),
            fourth_level=Value(fourth_level, output_field=DecimalField(max_digits=12, decimal_places=2)),
            avg_200=Window(expression=Avg('close'), order_by=F('date').asc(), frame=RowRange(start=-199, end=0)),
            avg_100=Window(expression=Avg('close'), order_by=F('date').asc(), frame=RowRange(start=-99, end=0)),
            avg_50=Window(expression=Avg('close'), order_by=F('date').asc(), frame=RowRange(start=-49, end=0)),
        ).order_by('date')
        # print(data)
        return data


class SectorAnalysis(UserAccessMixin, ListView):
    raise_exception = True
    permission_required = ('stocks.view_stockmap',)
    permission_denied_message = 'No permission'
    login_url = 'stocks:login'
    redirect_field_name = 'next'

    template_name = 'stocks/portfolio_stock_sector.html'
    context_object_name = 'stock_sector'

    def get_queryset(self):
        stockdata_subqs = StockData.objects.filter(company=OuterRef("pk")).values('company').alias(
            sumzero=Sum(F('quantity') * F('side'))).filter(
            sumzero__gt=0).annotate(price_sum=Sum(F('price') * F('side') * F('quantity')))

        stockmap_sector_sum_subqs = StockMap.objects.filter(sector=OuterRef("pk")).values('sector').annotate(
            comp_sec_sum=Sum(Subquery(stockdata_subqs.values('price_sum'))))

        sector_sum = Sector.objects.annotate(sector_sum=Sum((stockmap_sector_sum_subqs.values('comp_sec_sum')))).filter(
            sector_sum__gt=0)

        sector_total = 0
        for item in sector_sum:
            sector_total += item.sector_sum
        data = sector_sum.annotate(
            sector_total=Cast(sector_total, output_field=DecimalField(max_digits=12, decimal_places=2)),
            weight=F('sector_sum') * 100 / F('sector_total'))
        # print(data)
        return data


class MarketCapAnalysis(UserAccessMixin, ListView):
    permission_required = ('stocks.view_stockmap',)

    template_name = 'stocks/portfolio_stock_marketcap.html'
    context_object_name = 'stock_marketcap'

    def get_queryset(self):
        stockdata = StockData.objects.filter(company=OuterRef("pk")).values('company').alias(
            sumzero=Sum(F('quantity') * F('side'))).filter(
            sumzero__gt=0).annotate(price_sum=Sum(F('price') * F('side') * F('quantity')))
        stockmap_m_cap_sum = StockMap.objects.filter(m_cap=OuterRef("pk")).values('m_cap').annotate(
            comp_sec_sum=Sum(Subquery(stockdata.values('price_sum'))))
        m_cap_sum = MarketCap.objects.annotate(m_cap_sum=Sum((stockmap_m_cap_sum.values('comp_sec_sum')))).filter(
            m_cap_sum__gt=0)

        mcap_total = 0
        for item in m_cap_sum:
            mcap_total += item.m_cap_sum
        data = m_cap_sum.annotate(
            mcap_total=Cast(mcap_total, output_field=DecimalField(max_digits=12, decimal_places=2)),
            weight=F('m_cap_sum') * 100 / F('mcap_total'))

        # for i in data:
        #     print(i.__dict__)
        return data


class SearchView(UserAccessMixin, ListView):
    permission_required = ('stocks.view_stockmap',)

    template_name = 'stocks/search.html'
    context_object_name = 'search_results'

    # paginate_by = 9

    def get_queryset(self):

        search_text = self.request.GET.get('search')
        if search_text == 'new_closing_low':
            qsr = HistoricalData.objects.filter(company=OuterRef("pk"))

            get_data365 = qsr.filter(date__gte=date.today() - timedelta(days=365))
            today_hist = get_data365.order_by('-date')
            get_max = get_data365.order_by('-close')
            get_min = get_data365.order_by('close')
            get_avg = get_data365.values('company').annotate(average_price=Avg(F('close')))

            result = StockMap.objects.all().filter(is_portfolio_stock=True).annotate(
                max_close=Subquery(get_max.values('high')[:1]),
                max_date=Subquery(get_max.values('date')[:1]),
                min_close=Subquery(get_min.values('low')[:1]),
                min_date=Subquery(get_min.values('date')[:1]),
                close=Subquery(today_hist.values('close')[:1]),
                low=Subquery(today_hist.values('low')[:1]),
                high=Subquery(today_hist.values('high')[:1]),
                date=Subquery(today_hist.values('date')[:1]),
                avg_price=Subquery(get_avg.values('average_price')), ).filter(close__lte=F('min_close') * 1.04)
            return result

        elif search_text == 'new_closing_high':
            qsr = HistoricalData.objects.filter(company=OuterRef("pk"))

            get_data365 = qsr.filter(date__gte=date.today() - timedelta(days=365))
            today_hist = get_data365.order_by('-date')
            get_max = get_data365.order_by('-close')
            get_min = get_data365.order_by('close')
            get_avg = get_data365.values('company').annotate(average_price=Avg(F('close')))

            result = StockMap.objects.all().filter(is_portfolio_stock=True).annotate(
                max_close=Subquery(get_max.values('high')[:1]),
                max_date=Subquery(get_max.values('date')[:1]),
                min_close=Subquery(get_min.values('low')[:1]),
                min_date=Subquery(get_min.values('date')[:1]),
                close=Subquery(today_hist.values('close')[:1]),
                low=Subquery(today_hist.values('low')[:1]),
                high=Subquery(today_hist.values('high')[:1]),
                date=Subquery(today_hist.values('date')[:1]),
                avg_price=Subquery(get_avg.values('average_price')), ).filter(close__gte=F('max_close') * 1.01)
            return result


        else:
            result = StockMap.objects.order_by('name')
            # print(result)
            return result


class StockMapListView(UserAccessMixin, ListView):
    permission_required = ('stocks.view_stockmap',)

    model = StockMap
    template_name = 'stocks/stockmap_list.html'
    context_object_name = 'stockmap_list'
    queryset = StockMap.objects.order_by('name')


class StockMapDetailView(UserAccessMixin, DetailView):
    permission_required = ('stocks.view_stockmap',)

    model = StockMap
    template_name = 'stocks/stockmap_detail.html'
    # context_object_name = 'stockmapdetail'


class StockMapUpdateView(UserAccessMixin, UpdateView):
    permission_required = ('stocks.change_stockmap',)

    model = StockMap
    form_class = StockForm
    # fields = ['name','m_cap','sector','nse_symbol','moneycontrol_symbol',
    # 'yahoo_symbol','scrip_code','is_portfolio_stock']
    template_name_suffix = '_update_form'


class StockMapCreateView(UserAccessMixin, CreateView):
    permission_required = ('stocks.add_stockmap',)

    model = StockMap
    form_class = StockForm
    template_name = 'stocks/stockmap_add_form.html'


class StockMapDeleteView(UserAccessMixin, DeleteView):
    permission_required = ('stocks.delete_stockmap',)

    model = StockMap
    template_name = 'stocks/stockmap_delete.html'
    success_url = reverse_lazy('stocks:stockmap_list')

    def form_valid(self, form):
        success_url = self.get_success_url()
        self.object.soft_delete()
        return HttpResponseRedirect(success_url)


class SectorListView(UserAccessMixin, ListView):
    permission_required = ('stocks.view_sector',)

    model = Sector
    template_name = 'stocks/sector_list.html'
    context_object_name = 'sector_list'
    queryset = Sector.objects.order_by('sector')


class SectorDetailView(UserAccessMixin, DetailView):
    permission_required = ('stocks.view_sector',)

    model = Sector
    template_name = 'stocks/sector_detail.html'


class SectorUpdateView(UserAccessMixin, UpdateView):
    permission_required = ('stocks.change_sector',)

    model = Sector
    form_class = SectorForm
    # fields = ['name','m_cap','sector','nse_symbol','moneycontrol_symbol',
    # 'yahoo_symbol','scrip_code','is_portfolio_stock']
    template_name_suffix = '_update_form'


class SectorCreateView(UserAccessMixin, CreateView):
    permission_required = ('stocks.add_sector',)

    model = Sector
    form_class = SectorForm
    template_name = 'stocks/sector_add_form.html'


class SectorDeleteView(UserAccessMixin, DeleteView):
    permission_required = ('stocks.delete_sector',)

    model = Sector
    template_name = 'stocks/sector_delete.html'
    success_url = reverse_lazy('stocks:sector_list')

    def form_valid(self, form):
        success_url = self.get_success_url()
        self.object.soft_delete()
        return HttpResponseRedirect(success_url)


class StockDataListView(UserAccessMixin, ListView):
    permission_required = ('stocks.view_stockdata',)

    start_date = date.today()
    end_date = start_date - timedelta(days=365)

    model = StockData
    template_name = 'stocks/stockdata_list.html'
    context_object_name = 'stockdata_list'
    queryset = StockData.objects.filter(date__range=(end_date, start_date)).order_by('-date')


class StockDataDetailView(UserAccessMixin, DetailView):
    permission_required = ('stocks.view_stockdata',)

    model = StockData
    template_name = 'stocks/stockdata_detail.html'


class StockDataUpdateView(UserAccessMixin, UpdateView):
    permission_required = ('stocks.change_stockdata',)

    model = StockData
    form_class = StockDataForm
    template_name_suffix = '_update_form'


class StockDataCreateView(UserAccessMixin, CreateView):
    permission_required = ('stocks.add_stockdata',)

    model = StockData
    form_class = StockDataForm
    template_name = 'stocks/stockdata_add_form.html'


class StockDataDeleteView(UserAccessMixin, DeleteView):
    permission_required = ('stocks.delete_stockdata',)

    model = StockData
    template_name = 'stocks/stockdata_delete.html'
    success_url = reverse_lazy('stocks:stockdata_list')

    def form_valid(self, form):
        success_url = self.get_success_url()
        self.object.soft_delete()
        return HttpResponseRedirect(success_url)


class StockHistSearchView(ListView):
    template_name = 'stocks/historicaldata_list.html'
    context_object_name = 'historicaldata_list'

    def get_queryset(self):
        search_text = self.request.GET.get('name')
        qsr = HistoricalData.objects.order_by('-date').filter(company=search_text)[:300]
        return qsr

    def get_context_data(self, **kwargs):
        context = super(StockHistSearchView, self).get_context_data(**kwargs)
        context['form'] = StockFilterForm()
        return context


class HistoricalDataListView(UserAccessMixin, ListView):
    permission_required = ('stocks.view_historicaldata',)
    model = HistoricalData
    template_name = 'stocks/historicaldata_list.html'
    context_object_name = 'historicaldata_list'
    queryset = HistoricalData.objects.order_by('-date')[:300]

    # def get_queryset(self):
    #     search_text = self.request.GET.get('name')
    #     qsr = HistoricalData.objects.order_by('-date').filter(company=search_text)[:300]
    #     return qsr

    def get_context_data(self, **kwargs):
        context = super(HistoricalDataListView, self).get_context_data(**kwargs)
        context['form'] = StockFilterForm()
        return context


class HistoricalDataDetailView(UserAccessMixin, DetailView):
    permission_required = ('stocks.view_historicaldata',)

    model = HistoricalData
    template_name = 'stocks/historicaldata_detail.html'


class HistoricalDataUpdateView(UserAccessMixin, UpdateView):
    permission_required = ('stocks.change_historicaldata',)

    model = HistoricalData
    form_class = HistoricalDataForm
    # fields = ['name','m_cap','sector','nse_symbol','moneycontrol_symbol',
    # 'yahoo_symbol','scrip_code','is_portfolio_stock']
    template_name_suffix = '_update_form'


class HistoricalDataCreateView(UserAccessMixin, CreateView):
    permission_required = ('stocks.add_historicaldata',)

    model = HistoricalData
    form_class = HistoricalDataForm
    template_name = 'stocks/historicaldata_add_form.html'


class HistoricalDataDeleteView(UserAccessMixin, DeleteView):
    permission_required = ('stocks.delete_historicaldata',)

    model = HistoricalData
    template_name = 'stocks/historicaldata_delete.html'
    success_url = reverse_lazy('stocks:historicaldata_list')

    def form_valid(self, form):
        success_url = self.get_success_url()
        self.object.soft_delete()
        return HttpResponseRedirect(success_url)


def refresh_price_data(request):
    ticker_list = StockMap.objects.filter(is_portfolio_stock=True).order_by('id').values_list('id', 'nse_symbol')
    # print(ticker_list)
    for i, (id, comp) in enumerate(ticker_list):

        end_date = TODAY_DATE - timedelta(days=365)
        queryset = HistoricalData.objects.filter(
            Q(company=id) & Q(date__range=(end_date, TODAY_DATE))).order_by(
            'date')
        maximum_price = queryset.order_by('-high').first()
        minimum_price = queryset.order_by('low').first()
        prv_rec = queryset.order_by('-id')[:2:-1][:1]  # get last two record
        last_obj = queryset.last()
        print(prv_rec)
        change = last_obj.close - prv_rec[0].close
        perchange = change * 100 / prv_rec[0].close

        try:
            obj = NSEBhavcopy.objects.get(symbol=last_obj.company.id)
            obj.open = last_obj.open
            obj.high = last_obj.high
            obj.low = last_obj.low
            obj.close = last_obj.close
            obj.date = last_obj.date
            obj.adj_close = last_obj.adj_close
            obj.volume = last_obj.volume
            obj.w52_high = maximum_price.high
            obj.w52_low = minimum_price.low
            obj.prv_day_close = prv_rec[0].close
            obj.change = change
            obj.perchange = perchange
            obj.save()
            print(obj)

        except NSEBhavcopy.DoesNotExist:
            lastobj = NSEBhavcopy(
                is_deleted=False,
                deleted_at=None,
                symbol=StockMap.objects.get(nse_symbol=comp),
                date=last_obj.date,
                open=last_obj.open,
                high=last_obj.high,
                low=last_obj.low,
                close=last_obj.close,
                adj_close=last_obj.adj_close,
                volume=last_obj.volume,
                w52_high=maximum_price.high,
                w52_low=minimum_price.low,
                prv_day_close=prv_rec[0].close,
                change=change,
                perchange=perchange
            )
            lastobj.save()
            print(lastobj)

    return redirect(reverse('stocks:index'))


def BulkCreateHistoricalData(request):
    # ticker_list = StockMap.objects.filter(is_deleted=False).order_by('id').values_list('id', 'yahoo_symbol')

    ticker_list = StockMap.objects.filter(is_portfolio_stock=True).order_by('nse_symbol').values_list('id','nse_symbol', 'yahoo_symbol')
    for i, (id,nse_symbol, ticker) in enumerate(ticker_list):

        start_date = None
        end_date = None
        if TODAY_DATETIME < MARKET_CLOSE_DATETIME:
            end_date = TODAY_DATE - timedelta(days=1)
        else:
            end_date = TODAY_DATE

        last_obj = HistoricalData.objects.filter(company=id).order_by('date').last()
        if last_obj is None:
            start_date = date(2007, 1, 1)
            pass
        else:
            last_obj_date = last_obj.date.date()
            print(last_obj_date)
            day_in_word = last_obj.date.strftime("%A")
            print(day_in_word)
            if day_in_word == 'Friday' and last_obj_date == TODAY_DATE:
                print(f"1.Already Upto Date no need to update")
                pass
            elif day_in_word == 'Friday' and last_obj_date + timedelta(days=1) == TODAY_DATE:
                print(f"2.Already Upto Date no need to update")
                pass
            elif day_in_word == 'Friday' and last_obj_date + timedelta(days=2) == TODAY_DATE:
                print(f"3.Already Upto Date no need to update")
                pass
            else:
                start_date = last_obj_date + timedelta(days=1)

        if start_date >= end_date:
            print('start=', start_date, 'end=', end_date, 'Ticker=', ticker)
            print(f"4.Already Upto Date no need to update")
            pass

        else:
            # print(start_date)
            print('start=', start_date, 'end=', end_date)
            try:
                hist = yf.download(ticker, start=start_date, end=end_date,proxy=PROXIES)
            except:
                hist = yf.download(ticker, start=start_date, end=end_date)

            hist['date'] = hist.index
            hist['company_id'] = id
            hist.columns = map(str.lower, hist.columns)
            hist.columns = hist.columns.str.replace(' ','_')
            # hist['ticker'] = ticker
            hist = hist[['date', 'open', 'high', 'low', 'close', 'adj_close', 'volume', 'company_id']]
            hist = hist.dropna(
                subset=['date', 'open', 'high', 'low', 'close', 'adj_close', 'volume', 'company_id'])
            print(hist)
            datas = hist.to_records(index=False)

            bulkdata = []
            for data in datas:
                obj = HistoricalData(
                    is_deleted=False,
                    deleted_at=None,
                    date=pd.to_datetime(data.date),
                    open=data.open,
                    high=data.high,
                    low=data.low,
                    close=data.close,
                    adj_close=data.adj_close,
                    volume=int(data.volume),
                    company_id=data.company_id)
                bulkdata.append(obj)
                print(obj)
                # obj.save()
            HistoricalData.objects.bulk_create(bulkdata)
    return redirect(reverse('stocks:historicaldata_list'))


def str_to_float(string):
    bad_chars = [';', ':', '!', "*",
                 '?', '/', '₹', ',']
    test_string = string
    for i in bad_chars:
        test_string = test_string.replace(i, '')

    return float(test_string)


# def bulkCreateStockData(request):
#     excel_read_file = r'C:\Users\raj\Desktop\my_trade.xlsx'
#     df = pd.read_excel(excel_read_file, sheet_name="Data")
#     column = df.columns
#     df['date'] = pd.to_datetime(df['date']).dt.strftime('%Y/%d/%m')
#     df['price'] = df['price'].apply(str_to_float)
#     df = df.astype({'date': 'str', 'trade_time': 'str'})
#     df['DateAndTime'] = df['date'].str.cat(df['trade_time'], sep=" ")
#     df['date'] = pd.to_datetime(df['DateAndTime'])
#     df = df.drop(columns=['DateAndTime'])
#     df.sort_values("date", inplace=True)
#     df = df.reset_index(drop='index')
#     # print(df)
#     df = df.replace([np.nan], 0.0)
#     df['side'] = df['side'].map({'B': 1, 'S': -1, 'Buy': 1, 'Sell': -1})
#     df = df.astype(
#         {'date': 'datetime64[ns]', 'company': 'str', 'trade_num': 'int64', 'side': 'int', 'quantity': 'int',
#          'price': 'float'})
#     df = df.drop_duplicates(keep='first')
#     df = df[['date', 'company', 'side', 'quantity', 'price', 'trade_num']]
#     # print(df)
#     if df.empty:
#         print(f'There is no Data in {excel_read_file} for Upload in stocks_stockdata table')
#         ans = input(f'Enter any key for continue....... ')
#         pass
#     else:
#         df = df[['date', 'side', 'quantity', 'price', 'trade_num', 'company']]
#         print(df)
#         datas = df.values.tolist()
#         print(datas)
#         for data in datas:
#             obj = StockData(
#                 is_deleted=False,
#                 deleted_at=None,
#                 date=data[0],
#                 side=data[1],
#                 quantity=data[2],
#                 price=data[3],
#                 trade_num=data[4],
#                 company=StockMap.objects.get(name=data[5])
#             )
#             print(obj)
#             obj.save()
#         dfe = pd.DataFrame(columns=column)
#         outXLSX = pd.ExcelWriter(excel_read_file, engine='xlsxwriter')
#         dfe.to_excel(outXLSX, sheet_name='Data', index=False)
#         while True:
#             try:
#                 outXLSX.save()
#             except xlsxwriter.exceptions.FileCreateError as e:
#                 # For Python 3 use input() instead of raw_input().
#                 decision = input("Exception caught in workbook.close(): %s\n"
#                                  "Please close the file if it is open in Excel.\n"
#                                  "Try to write file again? [Y/n]: " % e)
#                 if decision != 'n':
#                     continue
#             break
#     return redirect(reverse('stocks:stockdata_list'))


def bulkCreateStockData(request):
    excel_read_file = r'C:\Users\raj\Desktop\my_trade.xlsx'
    # print(excel_read_file)
    workbook = openpyxl.load_workbook(filename=excel_read_file)
    sheet = workbook.active

    def side(data):
        data.strip()
        if data == 'Buy':
            return 1
        elif data == 'Sell':
            return -1

    for row in sheet.iter_rows(min_row=2, values_only=True):
        print(row)
        spread_date = row[0]
        spread_time = row[9]
        date_time = f'{spread_date} {spread_time}'
        # print(date_time)
        parsed_date = datetime.strptime(date_time, "%d-%m-%Y %H:%M:%S")
        # print(parsed_date)
        price = str_to_float(row[12].strip())
        comp_name = row[1].strip()
        print(comp_name)
        obj = StockData(
            is_deleted=False,
            deleted_at=None,
            date=parsed_date,
            side=side(row[10]),
            quantity=int(float(row[11])),
            price=price,
            trade_num=int(row[8]),
            company=StockMap.objects.get(name=comp_name)
        )
        print(obj.__dict__)
        obj.save()

    while (sheet.max_row > 1):
        # this method removes the row 2
        sheet.delete_rows(2)
    workbook.save(excel_read_file)

    return redirect(reverse('stocks:stockdata_list'))


def export_data_csv(request):
    buffer = io.BytesIO()
    wb = xlsxwriter.Workbook(buffer)
    ws = wb.add_worksheet()

    stockdata_subqs = StockData.objects.filter(company=OuterRef("pk")).filter(company__is_portfolio_stock=True).values(
        'company').alias(sumzero=Sum(F('quantity') * F('side'))).filter(
        sumzero__gt=0).annotate(qty_sum=Sum(F('quantity') * F('side')),
                                price_sum=Sum(F('price') * F('side') * F('quantity')),
                                average_price=Sum(F('price') * F('side') * F('quantity')) / Sum(
                                    F('quantity') * F('side')))

    nsedata_subqs = NSEBhavcopy.objects.filter(symbol=OuterRef("pk"))

    queryset = StockMap.objects.filter(is_portfolio_stock=True).annotate(
        qty_sum=Subquery(stockdata_subqs.values('qty_sum')),
        price_sum=Subquery(stockdata_subqs.values('price_sum')[:1]),
        average_price=Subquery(stockdata_subqs.values('average_price')[:1]),
        close=Subquery(nsedata_subqs.values('close')[:1]),
        date=Subquery(nsedata_subqs.values('date')[:1]),
        w52_high=Subquery(nsedata_subqs.values('w52_high')[:1]),
        w52_low=Subquery(nsedata_subqs.values('w52_low')[:1]),
        change=Subquery(nsedata_subqs.values('change')[:1]),
        pchange=Subquery(nsedata_subqs.values('perchange')[:1]),
        per_down=F('close') * 100 / F('w52_high') - 100,
        today_value=F('close') * F('qty_sum'),
        profit=F('today_value') - F('price_sum')
    ).order_by('name')
    # for item in data:
    #     print(item.name,item.profit)
    total_investment = 0
    profit_sum = 0

    for item in queryset:
        total_investment += item.price_sum
        profit_sum += item.profit

    queryset = queryset.annotate(
        total_investment=Cast(total_investment, output_field=DecimalField(max_digits=12, decimal_places=2)),
        profit_sum=Cast(profit_sum, output_field=DecimalField(max_digits=12, decimal_places=2)))
    # print(str(queryset.query))
    queryset = queryset.annotate(portweight=F('price_sum') * 100 / F('total_investment'),
                                 profit_percentage=F('today_value') * 100 / F('price_sum') - 100)

    datalist = queryset.values('name', 'nse_symbol', 'sector__sector', 'm_cap__name', 'qty_sum',
                               'average_price', 'price_sum', 'today_value', 'profit', 'close', 'w52_high', 'w52_low',
                               'change', 'pchange', 'per_down', 'profit_percentage', 'portweight', 'date')

    # print(datalist)
    row = 0
    col = 0

    bold_1 = wb.add_format({'bold': 1, 'font_size': 12})

    column_data = wb.add_format({'font_size': 11})
    date = wb.add_format({'font_size': 11, 'num_format': 'dd/mm/yyyy'})
    keys = datalist[0].keys()

    for key in keys:
        ws.write(row, col, key, bold_1)
        col += 1

    column_setting = [{'header': key} for key in keys]

    ws.add_table(0, 0, len(datalist), len(keys) - 1,
                 {'columns': column_setting, 'autofilter': True,
                  'style': 'Table Style Light 11'})
    col = 0
    row = 1

    for index, rows in enumerate(datalist):

        for i, (k, v) in enumerate(rows.items(), start=1):
            money = wb.add_format({'font_size': 11, 'num_format': '₹#,##0.00', 'font_color': 'black'})
            if col == 0:
                ws.write(row, col, v, bold_1)
            elif col > 4 and col <= 7:
                ws.write(row, col, v, money)
            elif col == 8:
                ws.write_datetime(row, col, v, date)
            elif col >= 9 and col <= 11:
                ws.write(row, col, v, money)
            elif col == 15:
                if v < 0:
                    money.set_font_color('red')
                    ws.write(row, col, v, money)
                else:
                    ws.write(row, col, v, money)
            else:
                ws.write(row, col, v, column_data)
            col += 1
        col = 0
        row += 1

    ws.write(row, col, "total investment:")
    col_1 = col + 1
    ws.write(row, col_1, f'=SUM(F2:F{row})')
    row_1 = row + 1
    ws.write(row_1, col_1 - 1, "today value:")
    ws.write(row_1, col_1, f'=SUM(O2:O{row})')
    row_2 = row_1 + 1
    ws.write(row_2, col_1 - 1, "total profit:")
    ws.write(row_2, col_1, f'=SUM(O2:O{row})-SUM(F2:F{row})')

    row_3 = row_2 + 1
    ws.write(row_3, col_1 - 1, "profit %:")
    ws.write(row_3, col_1, f'=(SUM(O2:O{row})-SUM(F2:F{row}))*100/SUM(F2:F{row})')

    column = [key for key in keys]
    length_list = [len(x) for x in column]

    for i, width in enumerate(length_list):
        ws.set_column(i, i, width)

    while True:
        try:
            wb.close()
        except xlsxwriter.exceptions.FileCreateError as e:
            decision = input("Exception caught in workbook.close(): %s\n"
                             "Please close the file if it is open in Excel.\n"
                             "Try to write file again? [Y/n]: " % e)
            if decision != 'n':
                continue
        break
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=True, filename=f'report{datetime.now()}.xlsx')
